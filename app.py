import os
import zipfile
import tempfile
import shutil
from openpyxl import load_workbook, Workbook
from flask import Flask, request, render_template, send_file, jsonify
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max upload size
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')  # Local uploads folder
app.config['ALLOWED_EXTENSIONS'] = {'xlsx', 'xls'}

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    """Check if the file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def split_excel(file_path, num_chunks):
    """Split an Excel file into the specified number of chunks"""
    # Load the Excel file
    wb = load_workbook(file_path)
    sheet = wb.active
    
    # Get the total number of rows (excluding header)
    total_rows = sheet.max_row - 1
    
    # Calculate rows per chunk (rounded up to ensure all rows are included)
    rows_per_chunk = -(-total_rows // num_chunks)  # Ceiling division
    
    # Get headers (first row)
    headers = [cell.value for cell in sheet[1]]
    
    # Create a temporary directory for the chunks
    chunk_dir = tempfile.mkdtemp()
    chunk_files = []
    
    # Split the workbook into chunks and save each chunk
    for i in range(num_chunks):
        start_idx = i * rows_per_chunk + 2  # +2 because Excel is 1-indexed and we skip header
        end_idx = min((i + 1) * rows_per_chunk + 1, total_rows + 1)
        
        # Skip if this chunk would be empty
        if start_idx > sheet.max_row:
            break
        
        # Create a new workbook for this chunk
        chunk_wb = Workbook()
        chunk_sheet = chunk_wb.active
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            chunk_sheet.cell(row=1, column=col_idx, value=header)
        
        # Copy data rows
        chunk_row = 2  # Start at row 2 (after header)
        for row_idx in range(start_idx, end_idx + 1):
            if row_idx <= sheet.max_row:  # Make sure we don't go beyond the actual data
                for col_idx in range(1, sheet.max_column + 1):
                    chunk_sheet.cell(row=chunk_row, column=col_idx, value=sheet.cell(row=row_idx, column=col_idx).value)
                chunk_row += 1
        
        # Create a filename for this chunk
        original_filename = os.path.basename(file_path)
        name_part = os.path.splitext(original_filename)[0]
        ext_part = os.path.splitext(original_filename)[1]
        chunk_filename = f"{name_part}_chunk_{i+1}{ext_part}"
        chunk_path = os.path.join(chunk_dir, chunk_filename)
        
        # Save the chunk as an Excel file
        chunk_wb.save(chunk_path)
        chunk_files.append(chunk_path)
    
    return chunk_dir, chunk_files

def create_zip(chunk_files):
    """Create a zip file containing all the chunk files"""
    # Create a temporary file for the zip
    zip_fd, zip_path = tempfile.mkstemp(suffix='.zip')
    os.close(zip_fd)
    
    # Create the zip file
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for chunk_file in chunk_files:
            zipf.write(chunk_file, os.path.basename(chunk_file))
    
    return zip_path

@app.route('/')
def index():
    """Render the main page"""
    return render_template('index.html')

@app.route('/split', methods=['POST'])
def split():
    """Handle file upload, splitting, and zip creation"""
    # Check if a file was uploaded
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    
    # Check if the file is empty
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    # Check if the file type is allowed
    if not allowed_file(file.filename):
        return jsonify({'error': 'File type not allowed'}), 400
    
    # Check if the number of chunks is valid
    try:
        num_chunks = int(request.form.get('chunks', '2'))
        if num_chunks < 2:
            return jsonify({'error': 'Number of chunks must be at least 2'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid number of chunks'}), 400
    
    try:
        # Save the uploaded file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Verify file exists after saving
        if not os.path.exists(file_path):
            return jsonify({'error': f'Failed to save uploaded file to {file_path}'}), 500
        
        # Split the Excel file
        chunk_dir, chunk_files = split_excel(file_path, num_chunks)
        
        # Create a zip file with the chunks
        zip_path = create_zip(chunk_files)
        
        # Clean up the temporary files
        try:
            os.remove(file_path)
            for chunk_file in chunk_files:
                if os.path.exists(chunk_file):
                    os.remove(chunk_file)
            if os.path.exists(chunk_dir):
                os.rmdir(chunk_dir)
        except Exception as e:
            app.logger.error(f"Error cleaning up files: {str(e)}")
        
        # Send the zip file as a download
        return send_file(
            zip_path,
            as_attachment=True,
            download_name=f"{os.path.splitext(filename)[0]}_chunks.zip",
            mimetype='application/zip'
        )
    
    except Exception as e:
        # Clean up any temporary files in case of an error
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        if 'chunk_dir' in locals() and os.path.exists(chunk_dir):
            shutil.rmtree(chunk_dir)
        if 'zip_path' in locals() and os.path.exists(zip_path):
            os.remove(zip_path)
        
        return jsonify({'error': str(e)}), 500

# Don't delete the uploads folder on teardown since it's a persistent directory
# Instead, we'll clean up individual files after processing

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
